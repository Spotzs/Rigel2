from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.conf import settings
from django.core.mail import send_mail
from django.db.models import Q
from django.urls import reverse_lazy
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
from django.http import HttpResponse, JsonResponse, HttpResponseRedirect
from django.template.loader import render_to_string
from django.urls import reverse
from django.contrib.auth.decorators import login_required, permission_required

from .forms import (
    ClienteForm,
    ProductoForm,
    ColegioForm,
    GradoForm,
    AniolectivoForm,
    ProveedorForm,
    PedidoForm,
    EmpleadoForm
)
from .models import Cliente, Producto, Colegio, Grado, Aniolectivo, Proveedor, Pedido, DetallePedido, Empleado
from django.views.generic import ListView, View

import csv
import io
import openpyxl

from reportlab.platypus import SimpleDocTemplate, Table, TableStyle
from reportlab.lib.pagesizes import letter
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet

from django.contrib.auth.decorators import user_passes_test
from django.shortcuts import redirect
from django.contrib import messages

@login_required
def exportar_listado(request):
    # Verificar si el usuario pertenece al grupo "Vendedor"
    if not request.user.groups.filter(name='Vendedor').exists():
       return redirect('/login/?denied=true')

    formato = request.GET.get('formato', 'csv')
    queryset = Producto.objects.all()

    if formato == 'csv':
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="productos.csv"'

        writer = csv.writer(response)
        writer.writerow(['Id_producto', 'Proveedor', 'Imagen', 'Nombre', 'Stock', 'Precio', 'Descripción', 'Fecha de Vencimiento'])

        for producto in queryset:
            writer.writerow([
                producto.Id_producto,
                producto.Proveedor,
                producto.Imagen_Producto.url if producto.Imagen_Producto else '',  # URL de la imagen si existe
                producto.Nombre,
                producto.Stock,
                producto.Precio,
                producto.Descripcion,
                producto.Fecha_vencimiento
            ])

        return response  # Devolver el objeto HttpResponse aquí

    elif formato == 'pdf':
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=letter, rightMargin=30, leftMargin=30, topMargin=30, bottomMargin=18)
        elements = []

        data = [['Id_producto', 'Proveedor', 'Nombre', 'Stock', 'Precio', 'Fecha de Vencimiento']]
        for producto in queryset:
            data.append([
                producto.Id_producto,
                str(producto.Proveedor),  # Convierte el objeto Proveedor a una cadena
                producto.Nombre,
                producto.Stock,
                producto.Precio,
                producto.Fecha_vencimiento
            ])

        table = Table(data)
        style = TableStyle([('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                            ('FONTSIZE', (0, 0), (-1, 0), 10),
                            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                            ('BOTTOMPADDING', (0, 1), (-1, -1), 6)])
        table.setStyle(style)
        elements.append(table)
        doc.build(elements)

        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = 'attachment; filename="productos.pdf"'
        response.write(buffer.getvalue())
        buffer.close()

        return response  # Devolver el objeto HttpResponse aquí

    elif formato == 'excel':
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.title = 'Productos'

        columns = ['Id_producto', 'Proveedor', 'Imagen', 'Nombre', 'Stock', 'Precio', 'Descripción', 'Fecha de Vencimiento']
        row_num = 1

        for col_num, column_title in enumerate(columns, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = column_title
            worksheet.column_dimensions[openpyxl.utils.get_column_letter(col_num)].width = 20  # Ajusta el ancho de la columna

        for producto in queryset:
            row_num += 1
            row = [
                producto.Id_producto,
                str(producto.Proveedor),  # Convertir el objeto Proveedor a una cadena de texto
                producto.Imagen_Producto.url if producto.Imagen_Producto else '',
                producto.Nombre,
                producto.Stock,
                producto.Precio,
                producto.Descripcion,
                producto.Fecha_vencimiento
            ]

            for col_num, cell_value in enumerate(row, 1):
                cell = worksheet.cell(row=row_num, column=col_num)
                cell.value = cell_value
                worksheet.row_dimensions[row_num].height = 30  # Ajusta el alto de la fila

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="productos.xlsx"'
        workbook.save(response)

        return response  # Devolver el objeto HttpResponse aquí

def paginate_queryset(queryset, page_number, items_per_page):
    paginator = Paginator(queryset, items_per_page)
    try:
        paginated_queryset = paginator.page(page_number)
    except PageNotAnInteger:
        paginated_queryset = paginator.page(1)
    except EmptyPage:
        paginated_queryset = paginator.page(paginator.num_pages)
    return paginated_queryset

# Vistas relacionadas con renderización de plantillas
def formularioContacto(request):
    return render(request, "formularioContacto.html")

@login_required
def Index(request):
    return render(request, "index.html")
def contactar(request):
    if request.method == "POST":
        asunto = request.POST["txtAsunto"]
        mensaje = f"{request.POST['txtMensaje']} / Email: {request.POST['txtEmail']}"
        email_desde = settings.EMAIL_HOST_USER
        email_para = ["zyaproyecto@gmail.com"]
        send_mail(asunto, mensaje, email_desde, email_para, fail_silently=False)
        return render(request, "contactoExitoso.html")
    return render(request, "formularioContacto.html")

@permission_required('Modulo_1.view_cliente', login_url='/login/')
def listado_cliente(request):
    clientes = Cliente.objects.all()
    registros_por_pagina = int(request.GET.get('registrosPorPagina', 5))  # Obtener el valor del parámetro GET, con 10 como valor predeterminado
    paginator = Paginator(clientes, registros_por_pagina)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    context = {
        'clientes': page_obj,
        'paginator': paginator,
        'page_obj': page_obj,
        'registros_por_pagina': registros_por_pagina,  # Agregar el valor de registros_por_pagina al contexto
    }

    return render(request, 'listclientes.html', context)

@permission_required('Modulo_1.view_colegio', login_url='/login/')
def listado_colegio(request):
    colegios = Colegio.objects.all()
    registros_por_pagina = int(request.GET.get('registrosPorPagina', 5))  # Obtener el valor del parámetro GET, con 10 como valor predeterminado
    paginator = Paginator(colegios, registros_por_pagina)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    context = {
        'colegios': page_obj,
        'paginator': paginator,
        'page_obj': page_obj,
        'registros_por_pagina': registros_por_pagina,  # Agregar el valor de registros_por_pagina al contexto
    }

    return render(request, 'listcole.html', context)


@permission_required('Modulo_1.view_grado', login_url='/login/')
def listado_grado(request):
    grados = Grado.objects.all()
    return render(request, "modals/Grados/listgrado.html", {'grados': grados})

@permission_required('Modulo_1.view_aniolectivo', login_url='/login/')
def listado_aniolectivo(request):
    anioslectivos = Aniolectivo.objects.all()
    return render(request, "modals/Lectivo/listlectivo.html", {'anioslectivos': anioslectivos})

@permission_required('Modulo_1.view_proveedor', login_url='/login/')
def listado_proveedor(request):
    proveedores = Proveedor.objects.all()
    registros_por_pagina = int(request.GET.get('registrosPorPagina', 5))  
    paginator = Paginator(proveedores, registros_por_pagina)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    context = {
        'proveedores': page_obj,
        'paginator': paginator,
        'page_obj': page_obj,
        'registros_por_pagina': registros_por_pagina,
    }

    return render(request, 'listproveedor.html', context)

@permission_required('Modulo_1.view_producto', login_url='/login/')
def listado_producto(request):
    productos = Producto.objects.all()
    registros_por_pagina = int(request.GET.get('registrosPorPagina', 5))  
    paginator = Paginator(productos, registros_por_pagina)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    context = {
        'productos': page_obj,
        'paginator': paginator,
        'page_obj': page_obj,
        'registros_por_pagina': registros_por_pagina,
    }

    return render(request, 'listproducto.html', context)

@permission_required('Modulo_1.view_pedido', login_url='/login/')
def listado_pedido(request):
    pedidos = Pedido.objects.all()
    registros_por_pagina = int(request.GET.get('registrosPorPagina', 5))  # Obtener el valor del parámetro GET, con 10 como valor predeterminado
    paginator = Paginator(pedidos, registros_por_pagina)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    context = {
        'pedidos': page_obj,
        'paginator': paginator,
        'page_obj': page_obj,
        'registros_por_pagina': registros_por_pagina,  # Agregar el valor de registros_por_pagina al contexto
    }

    return render(request, 'listpedidos.html', context)

@permission_required('Modulo_1.view_empleado', login_url='/login/')
def listado_empleado(request):
    empleados = Empleado.objects.all()
    registros_por_pagina = int(request.GET.get('registrosPorPagina', 5))  # Obtener el valor del parámetro GET, con 10 como valor predeterminado
    paginator = Paginator(empleados, registros_por_pagina)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    context = {
        'empleados': page_obj,
        'paginator': paginator,
        'page_obj': page_obj,
        'registros_por_pagina': registros_por_pagina,  # Agregar el valor de registros_por_pagina al contexto
    }

    return render(request, 'listempleados.html', context)

@permission_required('Modulo_1.add_estudiante', login_url='/login/')
def agregar_estudiante(request):
    if request.method == 'POST':
        form = ClienteForm(request.POST, files=request.FILES)
        if form.is_valid():
            form.save()
            messages.success(request, "Estudiante agregado correctamente")
            return redirect('listado_estudiante')
        else:
            messages.error(request, "Error al agregar el estudiante")
    else:
        form = ClienteForm()
    
    return render(request, 'agregarestudiante.html', {'form': form})

@login_required
def agregar_profesor(request):

    if not request.user.has_perm('Modulo_1.add_profesor'):
        return redirect('/login/?denied=true')

    current_page = request.META.get('HTTP_REFERER')  # Obtener la URL de la página actual

    if request.method == 'POST':
        form = ProfesorForm(request.POST, files=request.FILES)
        if form.is_valid():
            form.save()
            messages.success(request, "Profesor agregado correctamente")
            return redirect(current_page)
        else:
            messages.error(request, "Error al agregar el Profesor")
    else:
        form = ProfesorForm()

    return render(request, 'modals/Profesor/agregar_profesor.html', {'form': form})

@login_required
def agregar_empleado(request):

    if not request.user.has_perm('Modulo_1.add_empleado'):
        return redirect('/login/?denied=true')

    current_page = request.META.get('HTTP_REFERER')  # Obtener la URL de la página actual

    if request.method == 'POST':
        form = EmpleadoForm(request.POST, files=request.FILES)
        if form.is_valid():
            form.save()
            messages.success(request, "Empleado agregado correctamente")
            return redirect(current_page)
        else:
            messages.error(request, "Error al agregar el empleado")
    else:
        form = EmpleadoForm()

    return render(request, 'modals/Empleado/agregar_empleado.html', {'form': form})

@login_required
def agregar_producto(request):

    if not request.user.has_perm('Modulo_1.add_producto'):
        return redirect('/login/?denied=true')

    current_page = request.META.get('HTTP_REFERER')  # Obtener la URL de la página actual

    if request.method == 'POST':
        form = ProductoForm(request.POST, files=request.FILES)
        if form.is_valid():
            form.save()
            messages.success(request, "Producto agregado correctamente")
            return redirect(current_page)
        else:
            messages.error(request, "Error al agregar el producto")
    else:
        form = ProductoForm()

    return render(request, 'modals/Productos/agregar_producto.html', {'form': form})

@login_required
def agregar_colegio(request):

    if not request.user.has_perm('Modulo_1.add_colegio'):
        return redirect('/login/?denied=true')

    if request.method == 'POST':
        form = ColegioForm(request.POST, files=request.FILES)
        if form.is_valid():
            form.save()
            messages.success(request, "Colegio agregado correctamente")
            return redirect('listado_colegio')
        else:
            messages.error(request, "Error al agregar el colegio")
    else:
        form = ColegioForm()
    
    return render(request, 'modals/Colegio/agregar_colegio.html', {'form': form})

@login_required
def agregar_grado(request):

    if not request.user.has_perm('Modulo_1.add_grado'):
        return redirect('/login/?denied=true')

    if request.method == 'POST':
        form = GradoForm(request.POST, files=request.FILES)
        if form.is_valid():
            form.save()
            messages.success(request, "Grado agregado correctamente")
            return redirect('listado_grado')
        else:
            messages.error(request, "Error al agregar el grado")
    else:
        form = GradoForm()
    
    return render(request, 'modals/Grados/agregar_grado.html', {'form': form})

@login_required
def agregar_aniolectivo(request):

    if not request.user.has_perm('Modulo_1.add_aniolectivo'):
        return redirect('/login/?denied=true')

    if request.method == 'POST':
        form = AniolectivoForm(request.POST, files=request.FILES)
        if form.is_valid():
            form.save()
            messages.success(request, "Año lectivo agregado correctamente")
            return redirect('listado_aniolectivo')
        else:
            messages.error(request, "Error al agregar el Año")
    else:
        form = AniolectivoForm()
    
    return render(request, 'modals/Lectivo/agregar_aniolectivo.html', {'form': form})

@login_required
def agregar_proveedor(request):

    if not request.user.has_perm('Modulo_1.add_proveedor'):
        return redirect('/login/?denied=true')

    if request.method == 'POST':
        form = ProveedorForm(request.POST, files=request.FILES)
        if form.is_valid():
            form.save()
            messages.success(request, "Proveedor agregado correctamente")
            return redirect('listado_proveedor')
        else:
            messages.error(request, "Error al agregar el Proveedor")
    else:
        form = ProveedorForm()
    
    return render(request, 'modals/Proveedor/agregar_proveedor.html', {'form': form})

@login_required
def modificar_estudiante(request, Id_estudiante):

    if not request.user.has_perm('Modulo_1.change_estudiante'):
        return redirect('/login/?denied=true')

    estudiante = get_object_or_404(Estudiante, Id_estudiante=Id_estudiante)
    form = EstudianteForm(instance=estudiante)

    if request.method == 'POST':
        form = EstudianteForm(request.POST, instance=estudiante, files=request.FILES)
        if form.is_valid():
            form.save()
            messages.success(request, "Estudiante modificado correctamente")
            return redirect('listado_estudiante')
    
    return render(request, 'modificarestudiante.html', {'form': form})

@login_required
def modificar_empleado(request, Id_empleado):

    if not request.user.has_perm('Modulo_1.change_empleado'):
        return redirect('/login/?denied=true')

    empleado = get_object_or_404(Empleado, Id_empleado=Id_empleado)
    form = EmpleadoForm(instance=empleado)
    current_page = request.META.get('HTTP_REFERER')  # Obtener la URL de la página actual

    if request.method == 'POST':
        form = EmpleadoForm(request.POST, request.FILES, instance=empleado)
        if form.is_valid():
            form.save()
            messages.success(request, "Empleado modificado correctamente")
            return redirect(current_page)
        else:
            messages.error(request, "Error al modificar el empleado")

    return render(request, 'modals/Empleado/modificar_empleado.html', {'form': form, 'empleado': empleado})

@login_required
def modificar_grado(request, Codigo):

    if not request.user.has_perm('Modulo_1.change_grado'):
        return redirect('/login/?denied=true')

    grado = get_object_or_404(Grado, Codigo=Codigo)
    form = GradoForm(instance=grado)

    if request.method == 'POST':
        form = GradoForm(request.POST, instance=grado, files=request.FILES)
        if form.is_valid():
            form.save()
            messages.success(request, "Grado modificado correctamente")
            return redirect('listado_grado')
    
    return render(request, 'modals/Grados/modificar_grado.html', {'form': form})

@login_required
def modificar_colegio(request, Codigo):

    if not request.user.has_perm('Modulo_1.change_colegio'):
        return redirect('/login/?denied=true')

    colegio = get_object_or_404(Colegio, Codigo=Codigo)
    form = ColegioForm(instance=colegio)

    if request.method == 'POST':
        form = ColegioForm(request.POST, instance=colegio, files=request.FILES)
        if form.is_valid():
            form.save()
            messages.success(request, "Grado modificado correctamente")
            return redirect('listado_colegio')
    
    return render(request, 'modals/Colegio/modificar_colegio.html', {'form': form})

@login_required
def modificar_aniolectivo(request, Codigo):

    if not request.user.has_perm('Modulo_1.change_aniolectivo'):
        return redirect('/login/?denied=true')

    aniolectivo = get_object_or_404(Aniolectivo, Codigo=Codigo)
    form = AniolectivoForm(instance=aniolectivo)

    if request.method == 'POST':
        form = AniolectivoForm(request.POST, instance=aniolectivo, files=request.FILES)
        if form.is_valid():
            form.save()
            messages.success(request, "Año lectivo modificado correctamente")
            return redirect('listado_aniolectivo')
    
    return render(request, 'modals/Lectivo/modificar_aniolectivo.html', {'form': form})

# def modificar_aniolectivo(request, Codigo):
#     aniolectivo = get_object_or_404(Aniolectivo, Codigo=Codigo)
#     form = AniolectivoForm(instance=aniolectivo)

#     if request.method == 'POST':
#         form = AniolectivoForm(request.POST, instance=aniolectivo, files=request.FILES)
#         if form.is_valid():
#             form.save()
#             messages.success(request, "Año lectivo modificado correctamente")
#             return redirect('listado_Aniolectivo')
    
#     return render(request, 'modals/Colegio/modificar_colegio.html', {'form': form})

@login_required
def eliminar_estudiante(request, Id_estudiante):

    if not request.user.has_perm('Modulo_1.delete_estudiante'):
        return redirect('/login/?denied=true')

    estudiante = get_object_or_404(Estudiante, Id_estudiante=Id_estudiante)
    estudiante.delete()
    messages.success(request, "Estudiante eliminado correctamente")
    return redirect('listado_estudiante')

@login_required 
def eliminar_empleado(request, Id_empleado):

    if not request.user.has_perm('Modulo_1.delete_empleado'):
        return redirect('/login/?denied=true')

    empleado = get_object_or_404(Empleado, Id_empleado=Id_empleado)
    empleado_nombre = empleado.Nombres  # Guarda el nombre del producto antes de eliminarlo
    empleado.delete()
    # SweetAlert message
    swal_data = {
        'title': 'Empleado eliminado!',
        'text': f'El empleado "{empleado_nombre}" ha sido eliminado correctamente.',
        'icon': 'success',
    }
    # Devuelve la respuesta en formato JSON
    return JsonResponse(swal_data)

@login_required 
def eliminar_proveedor(request, pk):

    if not request.user.has_perm('Modulo_1.delete_proveedor'):
        return redirect('/login/?denied=true')

    proveedor = get_object_or_404(Proveedor, id=pk)
    proveedor_nombre = proveedor.nombre  
    proveedor.delete()

    swal_data = {
        'title': 'proveedor eliminado!',
        'text': f'El proveedor "{proveedor_nombre}" ha sido eliminado correctamente.',
        'icon': 'success',
    }

    return JsonResponse(swal_data)

def buscarestudiante(request):
    busqueda = request.GET.get("buscar")
    estudiantes = Estudiante.objects.all()

    if busqueda:
        estudiantes = estudiantes.filter(
            Q(Id_estudiante__icontains=busqueda) |
            Q(Apellido__icontains=busqueda) |
            Q(Nombres__icontains=busqueda) |
            Q(Genero__icontains=busqueda) |
            Q(Telefono__icontains=busqueda) |
            Q(Correo_electronico__icontains=busqueda) |
            Q(Direccion__icontains=busqueda) |
            Q(Activo__icontains=busqueda) |
            Q(Alergias__icontains=busqueda)
        ).distinct()

    return render(request, "listest.html", {"estudiantes": estudiantes})


@login_required
def buscar_grado(request):
    busqueda = request.GET.get("buscar")
    grados = Grado.objects.all()

    if busqueda:
        grados = grados.filter(
            Q(Codigo__icontains=busqueda) |
            Q(Nombre__icontains=busqueda) 
        ).distinct()

    return render(request, "modals/Grados/listgrado.html", {"grados": grados})


@login_required
def buscar_colegio(request):
    busqueda = request.GET.get("buscar")
    colegios = Colegio.objects.all()

    if busqueda:
        colegios = colegios.filter(
            Q(Codigo__icontains=busqueda) |
            Q(Nombre__icontains=busqueda) |
            Q(Rector__icontains=busqueda) 
        ).distinct()

    return render(request, "modals/Colegio/listcole.html", {"colegios": colegios})


@login_required
def buscar_lectivo(request):
    busqueda = request.GET.get("buscar")
    anioslectivos = Aniolectivo.objects.all()

    if busqueda:
        anioslectivos = anioslectivos.filter(
            Q(Codigo__icontains=busqueda) |
            Q(Nombre__icontains=busqueda) |
            Q(Fecha_inicio__icontains=busqueda) |
            Q(Fecha_fin__icontains=busqueda) 
        ).distinct()

    return render(request, "modals/Lectivo/listlectivo.html", {"anioslectivos": anioslectivos})

@login_required 
def eliminar_producto(request, Id_producto):

    if not request.user.has_perm('Modulo_1.delete_producto'):
        return redirect('/login/?denied=true')

    producto = get_object_or_404(Producto, Id_producto=Id_producto)
    producto_nombre = producto.Nombre  # Guarda el nombre del producto antes de eliminarlo
    producto.delete()
    # SweetAlert message
    swal_data = {
        'title': '¡Producto eliminado!',
        'text': f'El producto "{producto_nombre}" ha sido eliminado correctamente.',
        'icon': 'success',
    }
    # Devuelve la respuesta en formato JSON
    return JsonResponse(swal_data)

@login_required
def modificar_producto(request, pk):

    if not request.user.has_perm('Modulo_1.change_producto'):
        return redirect('/login/?denied=true')

    producto = get_object_or_404(Producto, pk=pk)
    form = ProductoForm(instance=producto)
    current_page = request.META.get('HTTP_REFERER')  # Obtener la URL de la página actual

    if request.method == 'POST':
        form = ProductoForm(request.POST, request.FILES, instance=producto)
        if form.is_valid():
            form.save()
            messages.success(request, "Producto modificado correctamente")
            return redirect(current_page)
        else:
            messages.error(request, "Error al modificar el producto")

    return render(request, 'modals/Productos/modificar_producto.html', {'form': form, 'producto': producto})

@login_required
def modificar_proveedor(request, pk):

    if not request.user.has_perm('Modulo_1.change_proveedor'):
        return redirect('/login/?denied=true')

    proveedor = get_object_or_404(Proveedor, pk=pk)
    form = ProveedorForm(instance=proveedor)
    current_page = request.META.get('HTTP_REFERER')  # Obtener la URL de la página actual

    if request.method == 'POST':
        form = ProveedorForm(request.POST, request.FILES, instance=proveedor)
        if form.is_valid():
            form.save()
            messages.success(request, "Proveedor modificado correctamente")
            return redirect(current_page)
        else:
            messages.error(request, "Error al modificar el Proveedor")

    return render(request, 'modals/Proveedor/modificar_proveedor.html', {'form': form, 'proveedor': proveedor})

@login_required
def buscar_producto(request):
    buscar = request.GET.get('buscar', '')
    productos_list = Producto.objects.all()

    # Capturar el número de página actual y los registros por página
    page_number = request.GET.get('page')
    registros_por_pagina = int(request.GET.get('registrosPorPagina', 5))  

    if buscar:
        productos_list = productos_list.filter(
            Q(Nombre__icontains=buscar) |
            Q(Descripcion__icontains=buscar) |
            Q(Stock__icontains=buscar)
        )

    paginator = Paginator(productos_list, registros_por_pagina)

    try:
        productos = paginator.page(page_number)
    except PageNotAnInteger:
        productos = paginator.page(1)
    except EmptyPage:
        productos = paginator.page(paginator.num_pages)

    context = {
        'productos': productos,
        'buscar': buscar,
        'registros_por_pagina': registros_por_pagina,
    }

    # Si la búsqueda está vacía, vuelve a cargar la misma página actual
    if not buscar.strip():
        # Construir la URL con los parámetros actuales
        url = reverse('listado_producto') + '?' + request.GET.urlencode()
        return redirect(url)

    return render(request, 'listproducto.html', context)


@login_required
def buscar_pedido(request):
    buscar = request.GET.get('buscar', '')
    pedidos_list = Pedido.objects.all()

    # Capturar el número de página actual y los registros por página
    page_number = request.GET.get('page')
    registros_por_pagina = int(request.GET.get('registrosPorPagina', 5))  

    if buscar:
        pedidos_list = pedidos_list.filter(
            Q(Nombre__icontains=buscar) |
            Q(Descripcion__icontains=buscar) |
            Q(Stock__icontains=buscar)
        )

    paginator = Paginator(pedidos_list, registros_por_pagina)

    try:
        pedidos = paginator.page(page_number)
    except PageNotAnInteger:
        pedidos = paginator.page(1)
    except EmptyPage:
        pedidos = paginator.page(paginator.num_pages)

    context = {
        'pedidos': pedidos,
        'buscar': buscar,
        'registros_por_pagina': registros_por_pagina,
    }

    # Si la búsqueda está vacía, vuelve a cargar la misma página actual
    if not buscar.strip():
        # Construir la URL con los parámetros actuales
        url = reverse('listado_pedido') + '?' + request.GET.urlencode()
        return redirect(url)

    return render(request, 'listpedidos.html', context)


def agregar_alcarrito(request):
    cliente_id = request.user.id
    cantidad = int(request.POST.get('cantidad'))
    producto_id = int(request.POST.get('producto_id'))
    producto = get_object_or_404(Producto, pk=producto_id)

    import pdb; pdb.set_trace(); print('Prueba 1')
    pedido = Pedido.objects.filter(
        estado='PENDIENTE', cliente_id=cliente_id).first()
    import pdb; pdb.set_trace(); print('Prueba 2')
    if not pedido:
        pedido = Pedido.objects.create(
            cliente=cliente_id,
            estado='PENDIENTE'
        )
    import pdb; pdb.set_trace(); print('Prueba 3')
    detalle_pedido = DetallePedido.objects.filter(
        pedido=pedido, producto=producto_id).first()
    precio_total = producto.precio * cantidad

    import pdb; pdb.set_trace(); print('Prueba 4')
    if not detalle_pedido:
        detalle_pedido = DetallePedido.objects.create(
            pedido=pedido,
            producto=producto_id,
            cantidad=cantidad,
            precio_total=precio_total
        )
    else:
        detalle_pedido.cantidad = cantidad
        detalle_pedido.precio_total = precio_total
        detalle_pedido.save()

    producto_nombre = producto.nombre  # Guarda el nombre del producto antes de eliminarlo
    # SweetAlert message
    swal_data = {
        'title': '¡Producto agregado al carrito!',
        'text': f'El producto "{producto_nombre}" ha sido agregado correctamente.',
        'icon': 'success',
    }
    # Devuelve la respuesta en formato JSON
    return JsonResponse(swal_data)